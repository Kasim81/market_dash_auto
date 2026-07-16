"""
sources/ons_rti.py
==================
ONS document-style XLSX datasets (§2.B B8) — e.g. PAYE Real Time Information
payrolled-employee counts, which are NOT exposed as a classic CDID timeseries
(so ``sources/ons.py`` can't carry them) but as a downloadable workbook behind
the ONS beta ``/data`` alias.

Two-step fetch:
  1. ``GET https://api.beta.ons.gov.uk/v1/data?uri=<dataset_uri>`` → JSON whose
     ``downloads[0].file`` is the current workbook filename (rotates monthly, so
     it must be read each run — the ``/current/`` alias in the uri is the stable
     part).
  2. ``GET https://www.ons.gov.uk/file?uri=<dataset_uri>/<filename>`` → XLSX.

Parse: a dates-DOWN sheet (month-name strings like "May 2026" in column A, a
value column identified by its header text). This is the opposite layout to the
dates-across ``sources/boe_survey.py``.

Series-id convention (a single library column):

    ``<dataset_uri>|<sheet>|<value_col_header>``

Indicator definitions live in ``data/macro_library_ons_rti.csv``.
"""

from __future__ import annotations

import datetime as _dt
import io
import json
import pathlib

import openpyxl
import pandas as pd

from sources.base import fetch_with_backoff

_LIBRARY_CSV = pathlib.Path(__file__).parent.parent / "data" / "macro_library_ons_rti.csv"
_BETA_DATA = "https://api.beta.ons.gov.uk/v1/data"
_ONS_FILE = "https://www.ons.gov.uk/file"
DEFAULT_HIST_START = "2014-01-01"
_UA = {"User-Agent": "Mozilla/5.0 (compatible; market-dash/1.0)"}


def load_library() -> list[dict]:
    """Load ONS-RTI indicator definitions from macro_library_ons_rti.csv."""
    if not _LIBRARY_CSV.exists():
        return []
    df = pd.read_csv(_LIBRARY_CSV, dtype=str, keep_default_na=False)
    if df.empty:
        return []
    df["sort_key"] = pd.to_numeric(df["sort_key"], errors="coerce").fillna(0)
    df = df.sort_values("sort_key")
    result = []
    for _, row in df.iterrows():
        result.append({
            "source":       "ONS RTI",
            "source_id":    row["series_id"].strip(),
            "col":          row["col"].strip(),
            "name":         row["name"].strip(),
            "country":      row.get("country", "").strip() or "GBR",
            "category":     row["category"].strip(),
            "subcategory":  row["subcategory"].strip(),
            "concept":      row.get("concept", "").strip(),
            "cycle_timing": row.get("cycle_timing", "").strip(),
            "units":        row["units"].strip(),
            "frequency":    row["frequency"].strip(),
            "notes":        row.get("notes", "").strip(),
            "sort_key":     float(row["sort_key"]),
            "series_id":    row["series_id"].strip(),
        })
    return result


def _resolve_workbook(dataset_uri: str, timeout: int = 45, retries: int = 3) -> bytes | None:
    """beta /data → current filename → download the XLSX bytes."""
    meta = fetch_with_backoff(
        _BETA_DATA, params={"uri": dataset_uri}, label=f"ONS RTI meta {dataset_uri}",
        retries=retries, timeout=timeout, headers=_UA,
    )
    if isinstance(meta, (bytes, bytearray)):
        try:
            meta = json.loads(meta)
        except ValueError:
            meta = None
    if not isinstance(meta, dict):
        print(f"    [ONS RTI] no metadata JSON for {dataset_uri}")
        return None
    downloads = meta.get("downloads") or []
    fname = next((d.get("file") for d in downloads if d.get("file", "").endswith((".xlsx", ".xls"))), None)
    if not fname:
        print(f"    [ONS RTI] no xlsx in downloads for {dataset_uri}: {downloads}")
        return None
    raw = fetch_with_backoff(
        _ONS_FILE, params={"uri": f"{dataset_uri}/{fname}"},
        label=f"ONS RTI {fname}", accept="bytes", retries=retries, timeout=timeout,
        headers=_UA,
    )
    return raw if isinstance(raw, (bytes, bytearray)) else None


def _parse_month(label) -> pd.Timestamp | None:
    s = str(label).strip()
    for fmt in ("%B %Y", "%b %Y", "%Y %b", "%Y-%m"):
        try:
            return pd.Timestamp(_dt.datetime.strptime(s, fmt).date().replace(day=1))
        except ValueError:
            continue
    return None


def parse_workbook(raw: bytes, series_id: str) -> pd.Series | None:
    """Parse ``<uri>|<sheet>|<value_col_header>`` from a dates-down ONS sheet."""
    try:
        _, sheet, value_header = series_id.split("|", 2)
    except ValueError:
        print(f"    [ONS RTI] bad series_id {series_id!r} (expected '<uri>|<sheet>|<value_header>')")
        return None
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    except Exception as e:                       # noqa: BLE001
        print(f"    [ONS RTI] workbook open failed for {series_id}: {e}")
        return None
    if sheet not in wb.sheetnames:
        print(f"    [ONS RTI] sheet {sheet!r} absent (have {wb.sheetnames[:8]})")
        return None
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))

    # header row: the one whose first cell is "Date"; the value column is the
    # cell in that row matching value_header (case-insensitive substring).
    header_idx = value_col = None
    vh = value_header.strip().lower()
    for i, r in enumerate(rows):
        if r and str(r[0]).strip().lower() == "date":
            for c, cell in enumerate(r):
                if cell is not None and vh in str(cell).strip().lower():
                    header_idx, value_col = i, c
                    break
            if header_idx is not None:
                break
    if header_idx is None:
        print(f"    [ONS RTI] header 'Date'/'{value_header}' not found in {sheet}")
        return None

    obs: dict[pd.Timestamp, float] = {}
    for r in rows[header_idx + 1:]:
        if not r or r[0] is None or value_col >= len(r) or r[value_col] is None:
            continue
        d = _parse_month(r[0])
        if d is None:
            continue
        try:
            obs[d] = float(r[value_col])
        except (ValueError, TypeError):
            continue
    if not obs:
        return None
    return pd.Series(obs).sort_index()


def fetch_series_as_pandas(
    series_id: str,
    col_name: str | None = None,
    start: str = DEFAULT_HIST_START,
) -> pd.Series | None:
    """Download + parse one ONS-RTI series. Signature matches the coordinator's
    ``_make_source_handlers`` factory (§2.C C2)."""
    dataset_uri = series_id.split("|", 1)[0]
    raw = _resolve_workbook(dataset_uri)
    if raw is None:
        return None
    s = parse_workbook(raw, series_id)
    if s is None or s.empty:
        return None
    s.name = col_name or series_id
    return s
