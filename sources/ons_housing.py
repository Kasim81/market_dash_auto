"""
sources/ons_housing.py
======================
ONS UK house-building workbook (§2.B B7) — permanent dwellings *started* and
*completed*, the free UK-wide successor to MHCLG's (discontinued-format) live
tables. "Dwellings started" is the standard leading construction-activity
indicator; MHCLG's own planning-permission live tables (Table 213) sit behind
``www.gov.uk``/``assets.publishing.service.gov.uk``, which are outside this
pipeline's egress policy, whereas the ONS release is on ``ons.gov.uk`` and is
reachable, UK-wide, and actively maintained (quarterly).

Two-step fetch, identical in shape to ``sources/ons_rti.py`` (B8):
  1. ``GET https://api.beta.ons.gov.uk/v1/data?uri=<dataset_uri>`` → JSON whose
     ``downloads[0].file`` is the current workbook filename. Here the filename
     is stable (``indicatorsofukhousebuilding.xlsx``) and the ``/current`` alias
     in the uri is the stable pointer, but we read it each run regardless.
  2. ``GET https://www.ons.gov.uk/file?uri=<dataset_uri>/<filename>`` → XLSX.

Parse: a dates-DOWN sheet, but unlike ONS RTI the date lives in the **Period**
column (a revision-flag column precedes it) and periods are quarter *ranges*
("Jan - Mar 1978"), not month names. The wanted value column is identified by
its header text (e.g. "Started - All Dwellings"). Suppression markers ("[x2]",
"[c]", …) in a cell are skipped.

Series-id convention (a single library column):

    ``<dataset_uri>|<sheet>|<value_col_header>``

Indicator definitions live in ``data/macro_library_ons_housing.csv``.
"""

from __future__ import annotations

import io
import json
import pathlib

import openpyxl
import pandas as pd

from sources.base import fetch_with_backoff

_LIBRARY_CSV = pathlib.Path(__file__).parent.parent / "data" / "macro_library_ons_housing.csv"
_BETA_DATA = "https://api.beta.ons.gov.uk/v1/data"
_ONS_FILE = "https://www.ons.gov.uk/file"
DEFAULT_HIST_START = "1978-01-01"
_UA = {"User-Agent": "Mozilla/5.0 (compatible; market-dash/1.0)"}

_MONTH_ABBR = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}


# ---------------------------------------------------------------------------
# LIBRARY LOADER
# ---------------------------------------------------------------------------

def load_library() -> list[dict]:
    """Load ONS-Housing indicator definitions from macro_library_ons_housing.csv."""
    if not _LIBRARY_CSV.exists():
        return []
    df = pd.read_csv(_LIBRARY_CSV, dtype=str, keep_default_na=False)
    if df.empty:
        return []
    df["sort_key"] = pd.to_numeric(df["sort_key"], errors="coerce").fillna(0)
    df = df.sort_values("sort_key")
    result = []
    for _, row in df.iterrows():
        result.append({
            "source":       "ONS Housing",
            "source_id":    row["series_id"].strip(),
            "col":          row["col"].strip(),
            "name":         row["name"].strip(),
            "country":      row.get("country", "").strip() or "GBR",
            "category":     row["category"].strip(),
            "subcategory":  row["subcategory"].strip(),
            "concept":      row.get("concept", "").strip(),
            "cycle_timing": row.get("cycle_timing", "").strip(),
            "units":        row["units"].strip(),
            "frequency":    row["frequency"].strip(),
            "notes":        row.get("notes", "").strip(),
            "sort_key":     float(row["sort_key"]),
            "series_id":    row["series_id"].strip(),
        })
    return result


# ---------------------------------------------------------------------------
# FETCH
# ---------------------------------------------------------------------------

def _resolve_workbook(dataset_uri: str, timeout: int = 45, retries: int = 3) -> bytes | None:
    """beta /data → current filename → download the XLSX bytes."""
    meta = fetch_with_backoff(
        _BETA_DATA, params={"uri": dataset_uri}, label=f"ONS Housing meta {dataset_uri}",
        retries=retries, timeout=timeout, headers=_UA,
    )
    if isinstance(meta, (bytes, bytearray)):
        try:
            meta = json.loads(meta)
        except ValueError:
            meta = None
    if not isinstance(meta, dict):
        print(f"    [ONS Housing] no metadata JSON for {dataset_uri}")
        return None
    downloads = meta.get("downloads") or []
    fname = next((d.get("file") for d in downloads if d.get("file", "").endswith((".xlsx", ".xls"))), None)
    if not fname:
        print(f"    [ONS Housing] no xlsx in downloads for {dataset_uri}: {downloads}")
        return None
    raw = fetch_with_backoff(
        _ONS_FILE, params={"uri": f"{dataset_uri}/{fname}"},
        label=f"ONS Housing {fname}", accept="bytes", retries=retries, timeout=timeout,
        headers=_UA,
    )
    return raw if isinstance(raw, (bytes, bytearray)) else None


# ---------------------------------------------------------------------------
# PARSER
# ---------------------------------------------------------------------------

def _parse_quarter(label) -> pd.Timestamp | None:
    """"Jan - Mar 1978" → quarter-end-month timestamp (1978-03-01). Returns None
    for annual/financial-year or non-quarter labels."""
    toks = str(label).replace("-", " ").split()
    year = next((int(t) for t in toks if t.isdigit() and len(t) == 4), None)
    month = next((_MONTH_ABBR[t[:3].lower()] for t in toks if t[:3].lower() in _MONTH_ABBR), None)
    if year is None or month is None:
        return None
    q_end_month = ((month - 1) // 3 + 1) * 3
    try:
        return pd.Timestamp(year, q_end_month, 1)
    except ValueError:
        return None


def parse_workbook(raw: bytes, series_id: str) -> pd.Series | None:
    """Parse ``<uri>|<sheet>|<value_col_header>`` from the dates-down ONS sheet."""
    try:
        _, sheet, value_header = series_id.split("|", 2)
    except ValueError:
        print(f"    [ONS Housing] bad series_id {series_id!r} "
              "(expected '<uri>|<sheet>|<value_header>')")
        return None
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    except Exception as e:                       # noqa: BLE001
        print(f"    [ONS Housing] workbook open failed for {series_id}: {e}")
        return None
    if sheet not in wb.sheetnames:
        print(f"    [ONS Housing] sheet {sheet!r} absent (have {wb.sheetnames[:8]})")
        wb.close()
        return None
    rows = list(wb[sheet].iter_rows(values_only=True))
    wb.close()

    # header row: the one carrying a "Period" cell; the date column is that
    # cell's index, the value column is the cell matching value_header (ci substr).
    header_idx = date_col = value_col = None
    vh = value_header.strip().lower()
    for i, r in enumerate(rows):
        if not r:
            continue
        pcol = next((c for c, cell in enumerate(r)
                     if cell is not None and str(cell).strip().lower() == "period"), None)
        if pcol is None:
            continue
        vcol = next((c for c, cell in enumerate(r)
                     if cell is not None and vh in str(cell).strip().lower()), None)
        if vcol is not None:
            header_idx, date_col, value_col = i, pcol, vcol
            break
    if header_idx is None:
        print(f"    [ONS Housing] header 'Period'/'{value_header}' not found in {sheet}")
        return None

    obs: dict[pd.Timestamp, float] = {}
    for r in rows[header_idx + 1:]:
        if not r or date_col >= len(r) or value_col >= len(r):
            continue
        d = _parse_quarter(r[date_col])
        if d is None:
            continue
        v = r[value_col]
        if v is None:
            continue
        try:
            obs[d] = float(v)           # suppression markers ("[x2]", "[c]") raise → skip
        except (ValueError, TypeError):
            continue
    if not obs:
        return None
    return pd.Series(obs).sort_index()


def fetch_series_as_pandas(
    series_id: str,
    col_name: str | None = None,
    start: str = DEFAULT_HIST_START,
) -> pd.Series | None:
    """Download + parse one ONS-Housing series. Signature matches the coordinator's
    ``_make_source_handlers`` factory (§2.C C2)."""
    dataset_uri = series_id.split("|", 1)[0]
    raw = _resolve_workbook(dataset_uri)
    if raw is None:
        return None
    s = parse_workbook(raw, series_id)
    if s is None or s.empty:
        return None
    s.name = col_name or series_id
    return s
