"""
sources/boe_survey.py
=====================
Bank of England *survey* spreadsheet source (§2.B B4 + B6).

The BoE publishes several quarterly surveys as XLSX workbooks — NOT as IADB
time-series codes — so they can't ride ``sources/boe.py``. This module downloads
one such workbook and pulls a single labelled row out of its wide (dates-across-
columns) layout. Two surveys are supported, keyed by the first ``|`` field of
the series id:

  * ``inflation-attitudes`` (B4) — a STATIC file under
    ``…/files/inflation-attitudes-survey/<stem>``; the date header row is real
    ``datetime`` cells. Series id::

        inflation-attitudes|<stem>|<sheet>|<question_substr>|<row_label>

  * ``credit-conditions`` (B6) — the Credit Conditions Survey annex, whose URL
    embeds the quarter (``…/credit-conditions-survey/<Y>/ccs-<Y>-q<N>-annex.xlsx``),
    so the workbook is resolved by walking back from the current quarter until a
    published annex is found. Its date header is a SPARSE year row + a quarter
    row (``2007`` … / ``Q2 Q3 Q4`` …). Series id::

        credit-conditions|<sheet>|<question_substr>|<row_label>

``<question_substr>`` matches (case-insensitive substring) the QUESTION header
cell in column A; ``<row_label>`` is the label — in column A (inflation) or
column B (credit-conditions sub-rows like "Past three months") — of the wanted
row at/after that question.

Retries via the shared ``sources.base.fetch_with_backoff`` engine (§2.C C3), in
mirror mode over the candidate URLs with an xlsx-magic-byte validator so a 404
on a not-yet-published quarter rotates to the next candidate.

Indicator definitions live in ``data/macro_library_boe_survey.csv``.
"""

from __future__ import annotations

import datetime as _dt
import io
import pathlib

import openpyxl
import pandas as pd

from sources.base import fetch_with_backoff

_LIBRARY_CSV = pathlib.Path(__file__).parent.parent / "data" / "macro_library_boe_survey.csv"
_BOE_MEDIA = "https://www.bankofengland.co.uk/-/media/boe/files"
DEFAULT_HIST_START = "1999-01-01"
_UA = {"User-Agent": "Mozilla/5.0 (compatible; market-dash/1.0)"}


# ---------------------------------------------------------------------------
# LIBRARY LOADER
# ---------------------------------------------------------------------------

def load_library() -> list[dict]:
    """Load BoE-survey indicator definitions from macro_library_boe_survey.csv."""
    if not _LIBRARY_CSV.exists():
        return []
    df = pd.read_csv(_LIBRARY_CSV, dtype=str, keep_default_na=False)
    if df.empty:
        return []
    df["sort_key"] = pd.to_numeric(df["sort_key"], errors="coerce").fillna(0)
    df = df.sort_values("sort_key")
    result = []
    for _, row in df.iterrows():
        result.append({
            "source":       "BoE Survey",
            "source_id":    row["series_id"].strip(),
            "col":          row["col"].strip(),
            "name":         row["name"].strip(),
            "country":      row.get("country", "").strip() or "GBR",
            "category":     row["category"].strip(),
            "subcategory":  row["subcategory"].strip(),
            "concept":      row.get("concept", "").strip(),
            "cycle_timing": row.get("cycle_timing", "").strip(),
            "units":        row["units"].strip(),
            "frequency":    row["frequency"].strip(),
            "notes":        row.get("notes", "").strip(),
            "sort_key":     float(row["sort_key"]),
            "series_id":    row["series_id"].strip(),
        })
    return result


# ---------------------------------------------------------------------------
# URL RESOLUTION
# ---------------------------------------------------------------------------

def _ccs_candidate_urls(n_back: int = 6) -> list[str]:
    """Latest-first candidate URLs for the Credit Conditions Survey annex,
    walking back ``n_back`` quarters from the current calendar quarter."""
    today = _dt.date.today()
    y, q = today.year, (today.month - 1) // 3 + 1
    urls = []
    for _ in range(n_back):
        urls.append(f"{_BOE_MEDIA}/credit-conditions-survey/{y}/ccs-{y}-q{q}-annex.xlsx")
        q -= 1
        if q == 0:
            q, y = 4, y - 1
    return urls


def _resolve_spec(series_id: str):
    """→ (candidate_urls, layout, sheet, question_substr, row_label) or None."""
    parts = series_id.split("|")
    survey = parts[0].strip()
    if survey == "inflation-attitudes" and len(parts) == 5:
        _, stem, sheet, question, row_label = parts
        return ([f"{_BOE_MEDIA}/inflation-attitudes-survey/{stem}"],
                "datetime_across", sheet, question, row_label)
    if survey == "credit-conditions" and len(parts) == 4:
        _, sheet, question, row_label = parts
        return (_ccs_candidate_urls(), "year_quarter_across", sheet, question, row_label)
    print(f"    [BoE Survey] unrecognised series_id {series_id!r}")
    return None


def _looks_xlsx(raw) -> str | None:
    """validate hook: accept only a real XLSX (zip) payload, reject 404/HTML."""
    if not isinstance(raw, (bytes, bytearray)) or raw[:2] != b"PK":
        return "not an xlsx (no PK header)"
    return None


def _download(urls: list[str], timeout: int = 45, retries: int = 3) -> bytes | None:
    raw = fetch_with_backoff(
        urls, label="BoE Survey", accept="bytes", retries=retries, timeout=timeout,
        headers=_UA, validate=_looks_xlsx,
    )
    return raw if isinstance(raw, (bytes, bytearray)) else None


# ---------------------------------------------------------------------------
# PARSERS
# ---------------------------------------------------------------------------

def _cell(row: tuple, c: int):
    return row[c] if row is not None and c < len(row) else None


def _find_target_row(rows: list, question_substr: str, row_label: str) -> int | None:
    """Index of the first row at/after the question whose label — in col A or
    col B — equals/startswith row_label (case-insensitive). Operates on a
    materialised row list (read-only ws.cell random access is pathologically
    slow on wide sheets, so callers pass ``list(ws.iter_rows(values_only=True))``)."""
    q_lower = question_substr.strip().lower()
    lbl = row_label.strip().lower()
    q_found = False
    for i, r in enumerate(rows):
        a = str(_cell(r, 0)).strip().lower() if _cell(r, 0) is not None else ""
        b = str(_cell(r, 1)).strip().lower() if _cell(r, 1) is not None else ""
        if not q_found:
            if a and q_lower in a:
                q_found = True     # fall through: the label may be on this row
            else:
                continue
        for cand in (a, b):
            if cand and (cand == lbl or cand.startswith(lbl)):
                return i
    return None


def _parse_datetime_across(rows: list, question, row_label) -> pd.Series | None:
    """Inflation-attitudes layout: one row of ``datetime`` header cells."""
    date_i = next((i for i, r in enumerate(rows[:40])
                   if sum(1 for c in (r[1:] if r else ()) if isinstance(c, _dt.datetime)) >= 3),
                  None)
    if date_i is None:
        return None
    tgt = _find_target_row(rows, question, row_label)
    if tgt is None:
        return None
    drow, trow = rows[date_i], rows[tgt]
    obs: dict[pd.Timestamp, float] = {}
    for c in range(1, min(len(drow), len(trow))):
        d, v = drow[c], trow[c]
        if isinstance(d, _dt.datetime) and v is not None:
            try:
                obs[pd.Timestamp(d.date())] = float(v)
            except (ValueError, TypeError):
                pass
    return pd.Series(obs).sort_index() if obs else None


def _parse_year_quarter_across(rows: list, question, row_label) -> pd.Series | None:
    """Credit-conditions layout: a sparse year row + a quarter row across
    columns (year printed once per year-block, forward-filled)."""
    year_i = qtr_i = None
    for i, r in enumerate(rows[:25]):
        vals = [str(x).strip() for x in (r or ()) if x is not None]
        if year_i is None and sum(1 for v in vals if v.isdigit() and len(v) == 4) >= 2:
            year_i = i
        if qtr_i is None and sum(1 for v in vals if v.upper() in ("Q1", "Q2", "Q3", "Q4")) >= 2:
            qtr_i = i
    if year_i is None or qtr_i is None:
        return None
    tgt = _find_target_row(rows, question, row_label)
    if tgt is None:
        return None
    yrow, qrow, trow = rows[year_i], rows[qtr_i], rows[tgt]
    obs: dict[pd.Timestamp, float] = {}
    cur_year = None
    for c in range(max(len(yrow), len(qrow), len(trow))):
        yv = _cell(yrow, c)
        if yv is not None and str(yv).strip().isdigit() and len(str(yv).strip()) == 4:
            cur_year = int(str(yv).strip())
        qv = _cell(qrow, c)
        q = str(qv).strip().upper() if qv is not None else ""
        v = _cell(trow, c)
        if cur_year is None or q not in ("Q1", "Q2", "Q3", "Q4") or v is None:
            continue
        try:
            obs[pd.Timestamp(cur_year, int(q[1]) * 3, 1)] = float(v)
        except (ValueError, TypeError):
            continue
    return pd.Series(obs).sort_index() if obs else None


def parse_workbook(raw: bytes, series_id: str) -> pd.Series | None:
    spec = _resolve_spec(series_id)
    if spec is None:
        return None
    _urls, layout, sheet, question, row_label = spec
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    except Exception as e:                       # noqa: BLE001
        print(f"    [BoE Survey] workbook open failed for {series_id}: {e}")
        return None
    match = next((s for s in wb.sheetnames if s.strip() == sheet.strip()), None)
    if match is None:
        print(f"    [BoE Survey] sheet {sheet!r} absent (have {wb.sheetnames})")
        wb.close()
        return None
    rows = list(wb[match].iter_rows(values_only=True))   # materialise once (read-only-safe)
    wb.close()
    if layout == "datetime_across":
        return _parse_datetime_across(rows, question, row_label)
    return _parse_year_quarter_across(rows, question, row_label)


def fetch_series_as_pandas(
    series_id: str,
    col_name: str | None = None,
    start: str = DEFAULT_HIST_START,
) -> pd.Series | None:
    """Download + parse one BoE-survey series. Signature matches the coordinator's
    ``_make_source_handlers`` factory (§2.C C2)."""
    spec = _resolve_spec(series_id)
    if spec is None:
        return None
    raw = _download(spec[0])
    if raw is None:
        return None
    s = parse_workbook(raw, series_id)
    if s is None or s.empty:
        return None
    s.name = col_name or series_id
    return s
