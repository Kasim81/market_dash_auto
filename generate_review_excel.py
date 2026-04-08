"""
generate_review_excel.py
========================
One-off script to generate indicator_groups_review.xlsx for user review.
Combines data from macro_indicator_library.csv with current hardcoded
group assignments from build_html.py.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Current hardcoded groups (from build_html.py lines 31-61) ────────────────
INDICATOR_GROUPS = {
    "US Growth & Style": [
        "US_G1","US_G2","US_G2b","US_G3","US_G3b","US_G4","US_G4b","US_G5","US_G6",
    ],
    "US Rates & Credit": [
        "US_I1","US_I2","US_I3","US_I4","US_I5","US_I6","US_I6b",
        "US_I7","US_I8","US_I9","US_I10","US_I11","US_R1","US_R2","US_RR1",
    ],
    "US FX & Momentum": [
        "US_FX1","US_FX2","M1","M2","M3","M4","M5",
    ],
    "US Macro Fundamentals": [
        "US_LEI1","US_JOBS1","US_LAB1","US_LAB2",
        "US_GROWTH1","US_HOUS1","US_M2L1","US_ISM1",
    ],
    "Europe & UK": [
        "EU_G1","EU_G2","EU_G3","EU_G4",
        "EU_I1","EU_I2","EU_I3","EU_I4","EU_R1","EU_FX1",
    ],
    "Asia: China & India": [
        "AS_G1","AS_G2","AS_G3","AS_G4",
        "AS_I1","AS_I2","AS_FX1","AS_FX2",
    ],
    "Asia Commodities & Japan": [
        "AS_C1","AS_C2","JP_G1","JP_FX1",
    ],
    "Global & Regional": [
        "REG_CLI1","REG_CLI2","REG_CLI3","REG_CLI4","REG_CLI5",
        "REG_RISK1","REG_EM1","REG_COMM1","REG_COMM2",
    ],
}

NATURALLY_LEADING = {
    "US_ISM1","US_LAB2","US_HOUS1","US_LEI1",
    "REG_CLI1","REG_CLI2","REG_CLI3","REG_CLI4","REG_CLI5","JP_FX1",
}

# ── Build reverse lookup: indicator_id → group_name ──────────────────────────
id_to_group = {}
for group_name, ids in INDICATOR_GROUPS.items():
    for ind_id in ids:
        id_to_group[ind_id] = group_name

# ── Load CSV ─────────────────────────────────────────────────────────────────
lib = pd.read_csv("data/macro_indicator_library.csv")

# ── Build rows in group order (matching current display order) ───────────────
rows = []
group_order = list(INDICATOR_GROUPS.keys())

for _, row in lib.iterrows():
    ind_id = str(row.get("id", "")).strip()
    if not ind_id:
        continue
    rows.append({
        "id": ind_id,
        "category": str(row.get("category", "")).strip(),
        "formula": str(row.get("formula_using_library_names", "")).strip(),
        "description": str(row.get("economic_interpretation", "")).strip(),
        "group": id_to_group.get(ind_id, "UNGROUPED"),
        "naturally_leading": "TRUE" if ind_id in NATURALLY_LEADING else "",
    })

# Sort by group order, then by position within group
def sort_key(r):
    group = r["group"]
    if group in group_order:
        g_idx = group_order.index(group)
        members = INDICATOR_GROUPS[group]
        m_idx = members.index(r["id"]) if r["id"] in members else 999
    else:
        g_idx = 999
        m_idx = 0
    return (g_idx, m_idx)

rows.sort(key=sort_key)

# ── Create Excel workbook ────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Indicator Groups Review"

# Header style
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# Data style
data_font = Font(name="Calibri", size=10)
wrap_align = Alignment(vertical="top", wrap_text=True)
group_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

headers = ["id", "category", "formula", "description", "group", "naturally_leading"]
col_widths = [12, 35, 55, 70, 28, 18]

# Write headers
for ci, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=ci, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# Write data
for ri, row_data in enumerate(rows, 2):
    for ci, key in enumerate(headers, 1):
        cell = ws.cell(row=ri, column=ci, value=row_data[key])
        cell.font = data_font
        cell.alignment = wrap_align
        cell.border = thin_border
        # Highlight the group column (editable)
        if key == "group":
            cell.fill = group_fill
            cell.font = Font(name="Calibri", size=10, bold=True)
        if key == "naturally_leading":
            cell.fill = group_fill

# Set column widths
for ci, w in enumerate(col_widths, 1):
    ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = w

# Freeze top row
ws.freeze_panes = "A2"

# Auto-filter
ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(headers)).column_letter}{len(rows)+1}"

output_path = "indicator_groups_review.xlsx"
wb.save(output_path)
print(f"Written {len(rows)} indicators to {output_path}")
print(f"Groups: {group_order}")
print(f"Naturally leading: {sorted(NATURALLY_LEADING)}")
