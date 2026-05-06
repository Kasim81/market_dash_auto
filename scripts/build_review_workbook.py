"""
Generates docs/dropdown_review.xlsx with two review tabs:

Tab 1  Macro Indicator Cleanup
       Maps every row of macro_indicator_library.csv to a proposed new
       sub_group value (collapsing the redundant "X - Y" pattern that
       duplicates information in subcategory).

Tab 2  Market Data Sub-Category Proposal
       Lists every market-data ticker with its full metadata and a
       proposed new Sub-Category that distinguishes Value/Growth/HY/IG/
       Energy/etc. instead of duplicating Broad Asset Class.

This is a review artefact only — nothing in this script edits the
source CSVs. Decisions made off the spreadsheet feed the next branch.
"""

import csv
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
IND_LIB = ROOT / "data" / "macro_indicator_library.csv"
MKT_HIST = ROOT / "data" / "market_data_comp_hist.csv"
OUT = ROOT / "docs" / "dropdown_review.xlsx"

# ── Tab 1: macro_indicator_library sub_group cleanup ─────────────────────────

SUB_GROUP_REMAP = {
    "Survey - PMI":                "Survey",
    "Survey - Business Sentiment": "Survey",
    "Survey - Sentiment":          "Survey",
    "Macro - Survey":              "Survey",
    "Mmtm - Credit":               "Momentum",
    "Mmtm - Equity":               "Momentum",
    "Mmtm - CrossAsset":           "Momentum",
    "Mmtm - Volatility":           "Momentum",
    "Growth Mmtm":                 "Momentum",
    "Equity - Factor (Size)":      "Equity Factor",
    "Equity - Factor (Style)":     "Equity Factor",
    "China - Equity - Factor (Size)": "Equity Factor",
    "India - Equity - Factor (Size)": "Equity Factor",
    "Equity - Growth":             "Equity",
    "China - Equity (Growth)":     "Equity",
    "CrossAsset - Growth":         "Cross-Asset",
    "CrossAsset - Inflation":      "Cross-Asset",
    "Rates - Growth":              "Rates",
    "Rates - Inflation":           "Rates",
    "China - Rates":               "Rates",
    "India - Rates":               "Rates",
    "China - FX Mmtm":             "FX",
    "India - FX Mmtm":             "FX",
    "EM - FX Mmtm":                "FX",
    "Japan - FX Mmtm":             "FX",
    "Growth (China broad)":        "Growth",
    "Growth (China infra)":        "Growth",
    # already-clean values left untouched (Macro, Credit, CLI, Volatility,
    # Inflation) are not in this map; column logic falls through to old.
}


def load_indicator_rows():
    with IND_LIB.open() as f:
        return list(csv.DictReader(f))


# ── Tab 2: market-data Sub-Category proposal ─────────────────────────────────

# Order matters: first matching rule wins.
EQUITY_RULES = [
    (re.compile(r"\bMin Vol|Low Vol|Low Volatility\b", re.I), "Low Volatility"),
    (re.compile(r"\bQuality\b", re.I),                         "Quality"),
    (re.compile(r"\bMomentum\b", re.I),                        "Momentum"),
    (re.compile(r"\bDividend\b", re.I),                        "Dividend"),
    (re.compile(r"\bValue\b", re.I),                           "Value"),
    (re.compile(r"\bGrowth\b", re.I),                          "Growth"),
    (re.compile(r"\bEqual[- ]?Weight", re.I),                  "Equal Weight"),
    (re.compile(r"\bSmall[- ]?Cap|Small Ordinaries|Smallcap|Russell 2000|FTSE SmallCap|S&P SmallCap|Nifty Smallcap\b", re.I), "Small Cap"),
    (re.compile(r"\bMid[- ]?Cap|Russell Midcap|Nifty Midcap|MDAX|S&P MidCap|FTSE 250|CAC Mid\b", re.I), "Mid Cap"),
    (re.compile(r"\bREIT|Real Estate\b", re.I),                "Sector — Real Estate"),
    (re.compile(r"\bFANG\b", re.I),                            "Thematic"),
]

# S&P 500 / STOXX / Global / Nifty sector ETFs — anything explicitly tagged
# with a sector name. The ticker namespace (^SP500-XXX, EXHx.DE, IX[A-Z],
# XL[A-Z], ^CNXxxx) is also a giveaway.
EQUITY_SECTOR_NAMES = {
    "Energy", "Materials", "Industrials", "Consumer Staples",
    "Consumer Discretionary", "Health Care", "Healthcare", "Financials",
    "Information Technology", "Tech", "Communication Services",
    "Comm Services", "Utilities", "Real Estate", "Banks", "Telecom",
    "Capital Goods", "Insurance", "Auto", "FMCG", "Pharma", "IT",
    "Metal", "Infrastructure",
}


def equity_sub(name: str, ticker: str) -> str:
    """Return proposed sub-category for an Equity ticker."""
    n = name or ""
    # Sector by ticker prefix (S&P sub-industries / S&P sectors / iShares globals)
    if ticker.startswith("^SP500-") or ticker.startswith("XL") and len(ticker) <= 4:
        return "Sector"
    if ticker.startswith("EXH") or ticker.startswith("EXV") or ticker.startswith("EXI5"):
        return "Sector"
    # Sector by name keyword
    for sector in EQUITY_SECTOR_NAMES:
        if re.search(rf"\b{re.escape(sector)}\b", n, re.I):
            # Avoid false-positive on "iShares Core MSCI World" etc. — these
            # don't contain sector words. The rule is: only treat as Sector
            # if the name *also* contains an index/ETF stem that's narrowly
            # sectoral. Heuristic: presence of a known sector word is enough.
            return "Sector"
    # Then style/factor rules
    for rx, lbl in EQUITY_RULES:
        if rx.search(n):
            return lbl
    return "Broad Market"


def bonds_sub(name: str, ticker: str, region: str, current_sub: str) -> str:
    if current_sub == "Rates":
        return "Government Yield"
    n = name or ""
    if re.search(r"\bInflation[- ]?Linked|Inflation[- ]?Protected|TIPS|Real Return\b", n, re.I):
        if re.search(r"\bEM|Emerging\b", n, re.I):
            return "EM Inflation-Linked"
        return "Inflation-Linked"
    is_em = bool(re.search(r"\bEM\b|Emerging|Argentina", n, re.I))
    is_hy = bool(re.search(r"\bHigh Yield|HY\b|BB|Single-B|CCC", n, re.I))
    is_corp = bool(re.search(r"\bCorp|Corporate", n, re.I))
    is_aggr = bool(re.search(r"\bAggregate\b", n, re.I))
    is_govt = bool(re.search(r"Government|Treasury|Gilt|Bund|OAT|BTP|CGB|JGB|Sovereign", n, re.I))
    if is_aggr:
        return "Aggregate"
    if is_em and is_hy:
        return "EM HY Credit"
    if is_em and is_corp:
        return "EM IG Credit"
    if is_em:
        return "EM Sovereign"
    if is_hy:
        return "HY Credit"
    if is_corp:
        return "IG Credit"
    if is_govt:
        return "DM Sovereign"
    return "DM Sovereign"


COMMODITY_BUCKETS = {
    "Energy":            re.compile(r"Crude|WTI|Brent|Gas|Heating|Gasoline|Energy", re.I),
    "Industrial Metals": re.compile(r"Copper|Aluminium|Aluminum|Iron Ore|Base Metals", re.I),
    "Precious Metals":   re.compile(r"Gold|Silver|Precious", re.I),
    "Agriculture":       re.compile(r"Corn|Wheat|Soybean|Sugar|Coffee|Cocoa|Cotton|Agriculture", re.I),
    "Livestock":         re.compile(r"Cattle|Hog|Lean", re.I),
    "Broad":             re.compile(r"All Commodities|Commodity Index|DB Commodity", re.I),
}


def commodity_sub(name: str) -> str:
    for bucket, rx in COMMODITY_BUCKETS.items():
        if rx.search(name or ""):
            return bucket
    return "Broad"


def fx_sub(name: str, ticker: str) -> str:
    n = name or ""
    if "DXY" in n or "Dollar Index" in n:
        return "DM Major"
    if re.search(r"\b(EUR|GBP|JPY|CHF|CAD|AUD|NZD|SEK|NOK)\b", n):
        return "DM Major"
    if re.search(r"\b(CNY|INR|KRW|TWD|MXN|BRL|TRY|ZAR|IDR|HKD)\b", n):
        return "EM Major"
    return "Other"


def crypto_sub(name: str) -> str:
    if re.search(r"Bitcoin|Ethereum", name or "", re.I):
        return "Major"
    return "Alt"


def vol_sub(name: str, ticker: str) -> str:
    n = name or ""
    if "MOVE" in n or "Bond Vol" in n:
        return "Bond Vol"
    if "OVX" in ticker or "Oil Vol" in n:
        return "Oil Vol"
    if "GVZ" in ticker or "Gold Vol" in n:
        return "Gold Vol"
    if "SKEW" in n:
        return "Equity Vol Tail"
    if "VXEFA" in ticker or "EAFE" in n:
        return "Equity Vol — Intl"
    if "VVIX" in n or "Volatility of VIX" in n:
        return "Equity Vol — Vol-of-Vol"
    return "Equity Vol"


def propose_market_sub(rec: dict) -> str:
    cls = rec.get("Broad Asset Class", "")
    nm = rec.get("Name", "")
    tk = rec.get("Ticker ID", "")
    reg = rec.get("Region", "")
    cur = rec.get("Sub-Category", "")
    if cls == "Equity":
        return equity_sub(nm, tk)
    if cls == "Bonds":
        return bonds_sub(nm, tk, reg, cur)
    if cls == "Commodities":
        return commodity_sub(nm)
    if cls == "FX":
        return fx_sub(nm, tk)
    if cls == "Crypto":
        return crypto_sub(nm)
    if cls == "Macro-Market Indicators":
        return vol_sub(nm, tk)
    return cur


def load_market_rows():
    with MKT_HIST.open() as f:
        r = csv.reader(f)
        rows = []
        for i, row in enumerate(r):
            rows.append(row)
            if i >= 10:
                break
    labels = [rr[1] for rr in rows]  # [Ticker ID, Variant, Source, Name, ...]
    seen = set()
    out = []
    for ci in range(2, len(rows[0])):
        rec = {lbl: rows[ri][ci].strip() for ri, lbl in enumerate(labels)}
        # Dedupe to one row per ticker — Local/USD variants share the
        # classification fields we care about for sub-category review.
        key = rec["Ticker ID"]
        if key in seen:
            continue
        seen.add(key)
        out.append(rec)
    return labels, out


# ── workbook assembly ────────────────────────────────────────────────────────

HDR_FILL = PatternFill("solid", fgColor="1F3A5F")
HDR_FONT = Font(color="FFFFFF", bold=True)
DIFF_FILL = PatternFill("solid", fgColor="FFF4CE")  # highlight changed rows


def write_tab1(ws):
    rows = load_indicator_rows()
    headers = [
        "id", "name (category)", "group",
        "old sub_group", "proposed sub_group",
        "subcategory", "concept", "cycle_timing",
    ]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
    n_changed = 0
    for r in rows:
        old_sg = r["sub_group"].strip()
        new_sg = SUB_GROUP_REMAP.get(old_sg, old_sg)
        ws.append([
            r["id"].strip(),
            r["category"].strip(),
            r["group"].strip(),
            old_sg,
            new_sg,
            r["subcategory"].strip(),
            r["concept"].strip(),
            r["cycle_timing"].strip(),
        ])
        if old_sg != new_sg:
            n_changed += 1
            for c in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=c).fill = DIFF_FILL
    # column widths
    widths = [22, 50, 20, 32, 22, 22, 22, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    print(f"  Tab 1: {len(rows)} indicator rows ({n_changed} would change sub_group)")


def write_tab2(ws):
    labels, rows = load_market_rows()
    extra = "Proposed Sub-Category"
    headers = list(labels) + [extra]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
    n_changed = 0
    for rec in rows:
        proposed = propose_market_sub(rec)
        ws.append([rec.get(lbl, "") for lbl in labels] + [proposed])
        if proposed != rec.get("Sub-Category", "").strip():
            n_changed += 1
            for c in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=c).fill = DIFF_FILL
    # column widths
    widths_by_label = {
        "Ticker ID": 18, "Variant": 9, "Source": 14, "Name": 56,
        "Broad Asset Class": 22, "Region": 18, "Sub-Category": 22,
        "Currency": 10, "Units": 10, "Frequency": 12, "Last Updated": 22,
    }
    for i, lbl in enumerate(labels, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths_by_label.get(lbl, 16)
    ws.column_dimensions[get_column_letter(len(labels) + 1)].width = 26
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    print(f"  Tab 2: {len(rows)} ticker/variant rows ({n_changed} proposed changes)")


def main():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Macro Indicator Cleanup"
    write_tab1(ws1)
    ws2 = wb.create_sheet("Market Data Sub-Category")
    write_tab2(ws2)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"\nWrote {OUT}")


if __name__ == "__main__":
    main()
